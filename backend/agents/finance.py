"""
EventOS Finance Agent — Budget Planner and Expense Tracker sub-agents.
"""

import os
import json
import asyncio
import uuid
from datetime import datetime
from backend.core.contracts import AgentLog, AgentResult
from backend.db import database

try:
    from google import genai
except ImportError:
    genai = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))


async def _emit(log_queue: asyncio.Queue, agent: str, msg: str, level: str = "info"):
    await log_queue.put(AgentLog(
        timestamp=datetime.utcnow(),
        agent_name=agent,
        domain="finance",
        message=msg,
        level=level,
    ))


async def budget_planner(params: dict, log_queue: asyncio.Queue, project_id: str = "default") -> AgentResult:
    """
    Budget Planner Sub-agent.
    Tools: Gemini 1.5 Flash for intelligent estimation + openpyxl for Excel export.
    """
    event_type = params.get("event_type", "hackathon")
    attendee_count = params.get("attendee_count", 500)
    duration_days = params.get("duration_days", 2)
    venue_cost = params.get("venue_cost", 0)

    await _emit(log_queue, "BUDGET_PLANNER",
                f"Planning budget for {event_type} — {attendee_count} attendees, {duration_days} days")

    categories = []
    total_budget = 0

    # Use Gemini for intelligent budget estimation
    api_key = os.getenv("GEMINI_API_KEY")

    if api_key and genai:
        try:
            client = genai.Client(api_key=api_key)

            prompt = f"""You are an expert event budget planner. Create a detailed budget breakdown.

Event: {event_type}
Attendees: {attendee_count}
Duration: {duration_days} day(s)
{"Known venue cost: $" + str(venue_cost) if venue_cost else "Venue cost: estimate based on event size"}

Create a comprehensive budget with these categories. Return ONLY a valid JSON object matching this schema exactly:
{{
    "total_budget": 0,
    "categories": [
        {{
            "name": "Category Name",
            "estimated": 0,
            "notes": "brief explanation",
            "subcategories": [{{"name": "...", "cost": 0}}]
        }}
    ]
}}

Required categories: Venue & Space, Catering & Food, AV Equipment & Tech, Marketing & Design, 
Speaker Fees & Travel, Swag & Merchandise, Insurance & Permits, Staffing & Volunteers, 
Prizes (if hackathon), Contingency (10% of total)."""

            response = client.models.generate_content(
                model="gemini-3.1-pro-preview",
                contents=prompt,
                config=genai.types.GenerateContentConfig(
                    temperature=0.3,
                    max_output_tokens=4096,
                    response_mime_type="application/json",
                ),
            )

            result = json.loads(response.text.strip())
            total_budget = result.get("total_budget", 0)

            for cat in result.get("categories", []):
                categories.append({
                    "name": cat.get("name", ""),
                    "estimated": cat.get("estimated", 0),
                    "actual": 0,
                    "notes": cat.get("notes", ""),
                    "subcategories": cat.get("subcategories", []),
                })
                await _emit(log_queue, "BUDGET_PLANNER",
                            f"  {cat.get('name', '')}: ${cat.get('estimated', 0):,.0f}")

        except Exception as e:
            await _emit(log_queue, "BUDGET_PLANNER", f"Gemini estimation error: {e}", "error")

    if not categories:
        # Fallback budget template
        await _emit(log_queue, "BUDGET_PLANNER", "Using fallback budget template", "warning")
        per_person = 50
        base = attendee_count * per_person
        categories = [
            {"name": "Venue & Space", "estimated": venue_cost or base * 0.25, "actual": 0, "notes": "Hall rental + setup", "subcategories": []},
            {"name": "Catering & Food", "estimated": base * 0.3, "actual": 0, "notes": f"${per_person * 0.3:.0f}/person", "subcategories": []},
            {"name": "AV Equipment & Tech", "estimated": base * 0.1, "actual": 0, "notes": "Sound, screens, streaming", "subcategories": []},
            {"name": "Marketing & Design", "estimated": base * 0.08, "actual": 0, "notes": "Digital + print materials", "subcategories": []},
            {"name": "Speaker Fees & Travel", "estimated": base * 0.12, "actual": 0, "notes": "Honorariums + travel", "subcategories": []},
            {"name": "Swag & Merchandise", "estimated": base * 0.05, "actual": 0, "notes": "T-shirts, stickers", "subcategories": []},
            {"name": "Insurance & Permits", "estimated": base * 0.03, "actual": 0, "notes": "Event insurance", "subcategories": []},
            {"name": "Staffing & Volunteers", "estimated": base * 0.02, "actual": 0, "notes": "Coordinator stipends", "subcategories": []},
            {"name": "Contingency (10%)", "estimated": base * 0.1, "actual": 0, "notes": "Buffer for unexpected costs", "subcategories": []},
        ]
        total_budget = sum(c["estimated"] for c in categories)

    await _emit(log_queue, "BUDGET_PLANNER",
                f"Budget generated — total estimate: ${total_budget:,.0f}", "success")

    # Generate Excel export
    excel_path = ""
    if openpyxl:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Event Budget"

            # Styling
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
            gold_font = Font(bold=True, size=14, color="DAA520")

            # Title
            ws["A1"] = f"EventOS Budget — {event_type.title()}"
            ws["A1"].font = Font(bold=True, size=16)
            ws["A2"] = f"Attendees: {attendee_count} | Duration: {duration_days} day(s)"
            ws["A2"].font = Font(size=10, color="888888")

            # Headers
            headers = ["Category", "Estimated ($)", "Actual ($)", "Remaining ($)", "% Spent", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data
            for i, cat in enumerate(categories, 5):
                ws.cell(row=i, column=1, value=cat["name"])
                ws.cell(row=i, column=2, value=cat["estimated"])
                ws.cell(row=i, column=3, value=cat["actual"])
                ws.cell(row=i, column=4, value=cat["estimated"] - cat["actual"])
                ws.cell(row=i, column=5, value="0%")
                ws.cell(row=i, column=6, value=cat["notes"])

            # Total row
            total_row = len(categories) + 5
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=total_budget).font = gold_font

            # Column widths
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 30

            excel_path = os.path.join(os.path.dirname(__file__), "..", "exports", f"budget_{project_id}.xlsx")
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            wb.save(excel_path)
            await _emit(log_queue, "BUDGET_PLANNER", f"Excel exported: {excel_path}", "success")

        except Exception as e:
            await _emit(log_queue, "BUDGET_PLANNER", f"Excel export failed: {e}", "warning")

    budget_data = {
        "project_id": project_id,
        "total_budget": total_budget,
        "total_spent": 0,
        "categories": categories,
        "excel_path": excel_path,
        "created_at": datetime.utcnow().isoformat(),
    }

    return AgentResult(
        agent_name="BUDGET_PLANNER",
        domain="finance",
        status="success",
        collection="budgets",
        data=budget_data,
    )


async def expense_tracker(params: dict, log_queue: asyncio.Queue, project_id: str = "default") -> AgentResult:
    """
    Expense Tracker Sub-agent.
    Logs expenses against budget categories and flags overruns.
    """
    expense = params.get("expense", {})
    category_name = expense.get("category", params.get("category", ""))
    amount = expense.get("amount", params.get("amount", 0))
    description = expense.get("description", params.get("description", ""))

    await _emit(log_queue, "EXPENSE_TRACKER",
                f"Logging expense: ${amount:,.2f} — {category_name}: {description}")

    # Read current budget from DB
    budget = await database.get_one_document("budgets", {"project_id": project_id})

    if not budget:
        await _emit(log_queue, "EXPENSE_TRACKER",
                    "No budget found — create a budget first with the Budget Planner", "error")
        return AgentResult(
            agent_name="EXPENSE_TRACKER",
            domain="finance",
            status="error",
            collection="budgets",
            data={"error": "No budget exists for this project"},
        )

    # Find the category and update
    categories = budget.get("categories", [])
    updated = False

    for cat in categories:
        if cat["name"].lower() == category_name.lower() or category_name.lower() in cat["name"].lower():
            cat["actual"] = cat.get("actual", 0) + amount
            spent_pct = (cat["actual"] / cat["estimated"] * 100) if cat["estimated"] > 0 else 100

            if spent_pct > 100:
                overrun = cat["actual"] - cat["estimated"]
                await _emit(log_queue, "EXPENSE_TRACKER",
                            f"🚨 OVERRUN: {cat['name']} exceeded budget by ${overrun:,.2f}", "error")
            elif spent_pct > 80:
                await _emit(log_queue, "EXPENSE_TRACKER",
                            f"⚠ WARNING: {cat['name']} at {spent_pct:.0f}% of budget", "warning")
            else:
                await _emit(log_queue, "EXPENSE_TRACKER",
                            f"✓ {cat['name']}: ${cat['actual']:,.2f} / ${cat['estimated']:,.2f} ({spent_pct:.0f}%)",
                            "success")

            updated = True
            break

    if not updated:
        await _emit(log_queue, "EXPENSE_TRACKER",
                    f"Category '{category_name}' not found in budget", "warning")
        categories.append({
            "name": category_name,
            "estimated": 0,
            "actual": amount,
            "notes": f"Auto-created for: {description}",
        })

    total_spent = sum(c.get("actual", 0) for c in categories)

    # Update in DB
    await database.update_document(
        "budgets",
        {"project_id": project_id},
        {"categories": categories, "total_spent": total_spent},
    )

    await _emit(log_queue, "EXPENSE_TRACKER",
                f"Expense logged — total spent: ${total_spent:,.2f}", "success")

    return AgentResult(
        agent_name="EXPENSE_TRACKER",
        domain="finance",
        status="success",
        collection="agent_logs",  # Don't duplicate-write to budgets since we already updated
        data={
            "action": "expense_logged",
            "category": category_name,
            "amount": amount,
            "description": description,
            "total_spent": total_spent,
            "project_id": project_id,
            "timestamp": datetime.utcnow().isoformat(),
        },
    )
